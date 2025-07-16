import openpyxl
from openpyxl import load_workbook

def listar_abas():
    try:
        yazo_wb = load_workbook('Yazo.xlsx', read_only=True)
        print("Abas disponíveis em Yazo.xlsx:")
        for i, sheet_name in enumerate(yazo_wb.sheetnames, 1):
            print(f"  {i}. '{sheet_name}' (comprimento: {len(sheet_name)})")
        return yazo_wb.sheetnames
    except FileNotFoundError:
        print('Arquivo Yazo.xlsx não encontrado!')
        return []
    except PermissionError:
        print('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
        return []

if __name__ == "__main__":
    listar_abas() 