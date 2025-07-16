import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

def verificar_celula_verde(cell):
    # Check if cell has the specific green background (RGB 0,255,0 / #00FF00)
    if cell.fill.start_color.rgb:
        rgb = cell.fill.start_color.rgb
        # Remove 'FF' prefix if present (Excel sometimes adds alpha channel)
        if rgb.startswith('FF'):
            rgb = rgb[2:]
        
        # Check for the specific green color used in the Excel filter
        if rgb.upper() == '00FF00':  # RGB(0,255,0) - Pure green
            return True
    
    return False

def verificar_linha_verde(worksheet, row_num):
    # Check if column A has green background
    cell = worksheet.cell(row=row_num, column=1)
    return verificar_celula_verde(cell)

def verificar_linha_verde_flexivel(worksheet, row_num):
    """Check if column A has green background"""
    # Check only column A (first column)
    cell = worksheet.cell(row=row_num, column=1)
    return verificar_celula_verde(cell)

def extrair_dados_plan(aba_nome):
    """Extract data from Plan.xlsx based on green background cells"""
    print(f'Lendo Plan.xlsx - aba {aba_nome}...')
    
    # Load workbook and worksheet
    plan_wb = load_workbook('Plan.xlsx')
    if aba_nome not in plan_wb.sheetnames:
        print(f'Aba {aba_nome} não encontrada em Plan.xlsx!')
        return []
    
    worksheet = plan_wb[aba_nome]
    linhas_extraidas = []
    linhas_verdes_encontradas = 0
    linhas_processadas = 0
    linhas_ignoradas = []
    
    # Process each row
    for row_num in range(2, worksheet.max_row + 1):  # Skip header row
        linhas_processadas += 1
        
        # Check if column A has green background
        cell_a = worksheet.cell(row=row_num, column=1)
        is_green = verificar_celula_verde(cell_a)
        
        if is_green:
            linhas_verdes_encontradas += 1
            
            # Extract data from specific columns based on mapping
            nome = worksheet.cell(row=row_num, column=1).value  # A
            local = worksheet.cell(row=row_num, column=14).value  # N
            titulo = worksheet.cell(row=row_num, column=6).value  # F
            descritivo = worksheet.cell(row=row_num, column=7).value  # G
            data = worksheet.cell(row=row_num, column=11).value  # K
            hora_inicio = worksheet.cell(row=row_num, column=12).value  # L
            hora_fim = worksheet.cell(row=row_num, column=13).value  # M
            
            # Check if we have at least a name or title
            if not nome and not titulo:
                linhas_ignoradas.append(f"Linha {row_num}: Sem nome e título")
                continue
            
            # Format data
            if data:
                try:
                    data_formatada = pd.to_datetime(str(data)).strftime('%Y/%m/%d')
                except:
                    data_formatada = str(data) if data else ''
            else:
                data_formatada = ''
            
            # Format time
            hora_inicio_str = str(hora_inicio) if hora_inicio else ''
            hora_fim_str = str(hora_fim) if hora_fim else ''
            
            # Determine tag based on aba
            tag = aba_nome.upper()
            
            linhas_extraidas.append({
                'ID Externo': '',
                'Local': str(local) if local else '',
                'Local (Inglês)': '',
                'Local (Espanhol)': '',
                'Título*': str(titulo) if titulo else '',
                'Título (Inglês)': '',
                'Título (Espanhol)': '',
                'Descrição': str(descritivo) if descritivo else '',
                'Descrição (Inglês)': '',
                'Descrição (Espanhol)': '',
                'Data (YYYY/MM/DD)*': data_formatada,
                'Horário de início (hh:mm:ss)*': hora_inicio_str,
                'Horário de término (hh:mm:ss)*': hora_fim_str,
                'Nome dos Palestrantes (separados por vírgula)': str(nome) if nome else '',
                'Tags (separadas por vírgulas)': tag,
                'Palavras-chave (separadas por vírgulas)': ''
            })
        else:
            # Debug: Check what color this cell actually has
            if cell_a.fill.start_color.rgb:
                rgb = cell_a.fill.start_color.rgb
                if rgb.startswith('FF'):
                    rgb = rgb[2:]
                linhas_ignoradas.append(f"Linha {row_num}: RGB={rgb}, Theme={cell_a.fill.start_color.theme}")
            elif cell_a.fill.fill_type and cell_a.fill.fill_type != 'none':
                linhas_ignoradas.append(f"Linha {row_num}: Fill type={cell_a.fill.fill_type}")
            else:
                linhas_ignoradas.append(f"Linha {row_num}: Sem preenchimento")
    
    print(f'Total de linhas processadas: {linhas_processadas}')
    print(f'Linhas com coluna A verde encontradas: {linhas_verdes_encontradas}')
    print(f'Linhas extraídas: {len(linhas_extraidas)}')
    
    # Show first few ignored lines for debugging
    if linhas_ignoradas:
        print(f'Primeiras 5 linhas ignoradas:')
        for i, linha in enumerate(linhas_ignoradas[:5]):
            print(f'  {linha}')
        if len(linhas_ignoradas) > 5:
            print(f'  ... e mais {len(linhas_ignoradas) - 5} linhas')
    
    return linhas_extraidas

def transferir_para_yazo(tipo_evento):
    """Transfer data to Yazo.xlsx based on event type"""
    mapeamento_abas = {
        '1': 'PALESTRA',
        '2': 'WORKSHOP', 
        '3': 'PAINEL'
    }
    
    aba_nome = mapeamento_abas.get(tipo_evento)
    if not aba_nome:
        print('Tipo de evento inválido!')
        return False
    
    # Extract data from Plan.xlsx
    linhas_extraidas = extrair_dados_plan(aba_nome)
    
    if not linhas_extraidas:
        print(f'Nenhuma linha com fundo verde encontrada na aba {aba_nome}!')
        return False
    
    print(f'Extraídas {len(linhas_extraidas)} linhas da aba {aba_nome}.')
    
    # Load Yazo.xlsx
    try:
        yazo_wb = load_workbook('Yazo.xlsx')
    except FileNotFoundError:
        print('Arquivo Yazo.xlsx não encontrado!')
        return False
    except PermissionError:
        print('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
        return False
    
    # Define columns for Yazo
    colunas_yazo = [
        'ID Externo', 'Local', 'Local (Inglês)', 'Local (Espanhol)', 'Título*', 'Título (Inglês)', 'Título (Espanhol)',
        'Descrição', 'Descrição (Inglês)', 'Descrição (Espanhol)', 'Data (YYYY/MM/DD)*', 'Horário de início (hh:mm:ss)*',
        'Horário de término (hh:mm:ss)*', 'Nome dos Palestrantes (separados por vírgula)', 'Tags (separadas por vírgulas)',
        'Palavras-chave (separadas por vírgulas)'
    ]
    
    # Try to find the correct sheet name - use Palestras2k25 as default
    target_sheet = None
    for sheet_name in yazo_wb.sheetnames:
        if 'Palestras2k25' in sheet_name or '2025' in sheet_name or sheet_name.strip() == '2025':
            target_sheet = sheet_name
            break
    
    # Check if target sheet exists, create if not
    if not target_sheet:
        target_sheet = 'Palestras2k25'
        yazo_wb.create_sheet(target_sheet)
        # Add headers
        ws = yazo_wb[target_sheet]
        for col, header in enumerate(colunas_yazo, 1):
            ws.cell(row=1, column=col, value=header)
    
    print(f'Usando aba: "{target_sheet}"')
    
    # Read existing data
    try:
        yazo_df = pd.read_excel('Yazo.xlsx', sheet_name=target_sheet)
    except:
        # Create empty DataFrame with correct columns
        yazo_df = pd.DataFrame({col: pd.Series(dtype='str') for col in colunas_yazo})
    
    # Add new rows
    novo_df = pd.DataFrame(linhas_extraidas)
    resultado_df = pd.concat([yazo_df, novo_df], ignore_index=True)
    
    # Save back to Yazo.xlsx
    try:
        with pd.ExcelWriter('Yazo.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            resultado_df.to_excel(writer, sheet_name=target_sheet, index=False)
    except PermissionError:
        print('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
        return False
    
    print(f'Transferência concluída! {len(linhas_extraidas)} registros adicionados à aba "{target_sheet}".')
    return True

def listar_abas_yazo():
    """List all available sheets in Yazo.xlsx"""
    try:
        yazo_wb = load_workbook('Yazo.xlsx')
        print("Abas disponíveis em Yazo.xlsx:")
        for i, sheet_name in enumerate(yazo_wb.sheetnames, 1):
            print(f"  {i}. '{sheet_name}'")
        return yazo_wb.sheetnames
    except FileNotFoundError:
        print('Arquivo Yazo.xlsx não encontrado!')
        return []
    except PermissionError:
        print('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
        return []

def apagar_dados_yazo():
    """Clear all data below header in Yazo.xlsx sheet"""
    try:
        yazo_wb = load_workbook('Yazo.xlsx')
    except FileNotFoundError:
        print('Arquivo Yazo.xlsx não encontrado!')
        return False
    except PermissionError:
        print('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
        return False
    
    # List available sheets first
    print("Abas disponíveis em Yazo.xlsx:")
    for sheet_name in yazo_wb.sheetnames:
        print(f"  - '{sheet_name}'")
    
    # Try to find the correct sheet name - use Palestras2k25 as default
    target_sheet = None
    for sheet_name in yazo_wb.sheetnames:
        if 'Palestras2k25' in sheet_name or '2025' in sheet_name or sheet_name.strip() == '2025':
            target_sheet = sheet_name
            break
    
    if not target_sheet:
        print('Aba Palestras2k25 ou 2025 não encontrada em Yazo.xlsx!')
        print('Abas disponíveis:', yazo_wb.sheetnames)
        return False
    
    print(f'Usando aba: "{target_sheet}"')
    
    # Get the worksheet
    ws = yazo_wb[target_sheet]
    
    # Clear all rows except header (row 1)
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    
    # Save the workbook
    yazo_wb.save('Yazo.xlsx')
    print(f'Todos os dados da aba "{target_sheet}" foram apagados (mantendo apenas o cabeçalho).')
    return True

def main():
    opcoes_utilizadas = set()  # Track used options in current session
    
    while True:
        print("\n=== PROGRAMA DE TRANSFERÊNCIA DE DADOS ===")
        print("Opções de importação para Yazo:")
        print("1 - Palestra")
        print("2 - Workshop") 
        print("3 - Painel")
        print("4 - Fechar")
        print("5 - Apagar tudo")
        
        escolha = input("\nEscolha uma opção (1-5): ").strip()
        
        if escolha == "4":
            print("Encerrando programa...")
            break
            
        elif escolha == "5":
            if escolha in opcoes_utilizadas:
                confirmacao = input("Você já utilizou esta opção nesta sessão. Deseja prosseguir? (s/n): ").strip().lower()
                if confirmacao != 's':
                    continue
            
            confirmacao = input("ATENÇÃO: Esta ação irá APAGAR TODOS os registros da aba Palestras2k25. Deseja prosseguir? (s/n): ").strip().lower()
            if confirmacao == 's':
                apagar_dados_yazo()
                opcoes_utilizadas.add(escolha)
            else:
                print("Operação cancelada.")
                
        elif escolha in ["1", "2", "3"]:
            if escolha in opcoes_utilizadas:
                confirmacao = input(f"Você já utilizou a opção {escolha} nesta sessão. Deseja prosseguir? (s/n): ").strip().lower()
                if confirmacao != 's':
                    continue
            
            if transferir_para_yazo(escolha):
                opcoes_utilizadas.add(escolha)
                
        else:
            print("Opção inválida! Escolha uma opção de 1 a 5.")

if __name__ == "__main__":
    main() 