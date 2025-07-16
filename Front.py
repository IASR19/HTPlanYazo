import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os

class DataTransferApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PlanYazo - Transferência de Dados")
        self.root.geometry("1200x800")  # Aumentado ainda mais o tamanho da janela
        self.root.configure(bg='#f8f9fa')
        
        # Set app icon and style
        self.setup_styles()
        self.create_widgets()
        
        # Track used options
        self.opcoes_utilizadas = set()
        
        # Initial file check after all widgets are created
        self.root.after(100, self.check_files)
        
    def setup_styles(self):
        """Configure modern styling for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        style.configure('Title.TLabel', font=('Arial', 28, 'bold'), foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Arial', 18, 'bold'), foreground='#34495e')
        style.configure('Info.TLabel', font=('Arial', 14), foreground='#7f8c8d')
        
        # Button styles
        style.configure('Primary.TButton', 
                       font=('Arial', 14, 'bold'),
                       background='#3498db',
                       foreground='white')
        style.configure('Success.TButton',
                       font=('Arial', 14, 'bold'),
                       background='#27ae60',
                       foreground='white')
        style.configure('Danger.TButton',
                       font=('Arial', 14, 'bold'),
                       background='#e74c3c',
                       foreground='white')
        style.configure('Info.TButton',
                       font=('Arial', 12),
                       background='#95a5a6',
                       foreground='white')
        
    def create_widgets(self):
        """Create and arrange all GUI widgets"""
        # Main container with proper spacing
        main_frame = ttk.Frame(self.root, padding="30")  # Reduzido padding para caber melhor
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)  # Make log section expandable
        
        # Create sections in proper order with more spacing
        self.create_title_section(main_frame, 0)
        self.create_file_status_section(main_frame, 1)
        self.create_transfer_section(main_frame, 2)
        self.create_log_section(main_frame, 3)
        
    def create_title_section(self, parent, row):
        """Create title section with proper spacing"""
        title_frame = ttk.Frame(parent)
        title_frame.grid(row=row, column=0, sticky="ew", pady=(0, 20))  # Reduzido espaçamento
        title_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(title_frame, text="PlanYazo", style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 10))  # Reduzido espaçamento
        
        subtitle_label = ttk.Label(title_frame, text="Sistema de Transferência de Dados", style='Info.TLabel')
        subtitle_label.grid(row=1, column=0)
        
    def create_file_status_section(self, parent, row):
        """Create file status display section with proper spacing"""
        # Section header
        header_label = ttk.Label(parent, text="Status dos Arquivos", style='Header.TLabel')
        header_label.grid(row=row, column=0, sticky="w", pady=(0, 15))  # Reduzido espaçamento
        
        # File status frame
        status_frame = ttk.Frame(parent)
        status_frame.grid(row=row+1, column=0, sticky="ew", pady=(0, 20))  # Reduzido espaçamento
        status_frame.columnconfigure(1, weight=1)
        
        # Plan.xlsx status
        ttk.Label(status_frame, text="Plan.xlsx:", style='Info.TLabel').grid(row=0, column=0, sticky="w", padx=(0, 20), pady=10)  # Reduzido espaçamento
        self.plan_status = ttk.Label(status_frame, text="Verificando...", style='Info.TLabel')
        self.plan_status.grid(row=0, column=1, sticky="w", pady=10)
        
        # Yazo.xlsx status
        ttk.Label(status_frame, text="Yazo.xlsx:", style='Info.TLabel').grid(row=1, column=0, sticky="w", padx=(0, 20), pady=10)  # Reduzido espaçamento
        self.yazo_status = ttk.Label(status_frame, text="Verificando...", style='Info.TLabel')
        self.yazo_status.grid(row=1, column=1, sticky="w", pady=10)
        
        # Refresh button
        refresh_btn = ttk.Button(status_frame, text="🔄 Atualizar", 
                                command=self.check_files, style='Primary.TButton')
        refresh_btn.grid(row=0, column=2, rowspan=2, padx=(50, 0), pady=10)  # Reduzido espaçamento
        
    def create_transfer_section(self, parent, row):
        """Create transfer options section with proper spacing"""
        # Section header
        
        # Transfer buttons frame - reorganizado para evitar cortes
        transfer_frame = ttk.Frame(parent)
        transfer_frame.grid(row=row+1, column=0, sticky="ew", pady=(0, 20))  # Reduzido espaçamento
        transfer_frame.columnconfigure(4, weight=1)  # Push clear button to the right
        
        # Transfer buttons with much more spacing
        self.palestra_btn = ttk.Button(transfer_frame, text="🎤 Palestra", 
                                      command=lambda: self.transfer_data('1'), style='Primary.TButton')
        self.palestra_btn.grid(row=0, column=0, padx=(0, 20), pady=15)  # Reduzido espaçamento
        
        self.workshop_btn = ttk.Button(transfer_frame, text="🔧 Workshop", 
                                      command=lambda: self.transfer_data('2'), style='Primary.TButton')
        self.workshop_btn.grid(row=0, column=1, padx=(0, 20), pady=15)  # Reduzido espaçamento
        
        self.painel_btn = ttk.Button(transfer_frame, text="👥 Painel", 
                                    command=lambda: self.transfer_data('3'), style='Primary.TButton')
        self.painel_btn.grid(row=0, column=2, padx=(0, 30), pady=15)  # Reduzido espaçamento
        
        # Clear data button
        self.clear_btn = ttk.Button(transfer_frame, text="🗑️ Apagar Dados", 
                                   command=self.clear_data, style='Danger.TButton')
        self.clear_btn.grid(row=0, column=3, padx=(0, 0), pady=15, sticky="e")  # Reduzido espaçamento
        
    def create_log_section(self, parent, row):
        """Create log display section with proper spacing"""
        # Section header
        # Log text area with scrollbar
        log_frame = ttk.Frame(parent)
        log_frame.grid(row=row+1, column=0, sticky="nsew", pady=(0, 10))  # Reduzido espaçamento
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Text widget with scrollbar
        self.log_text = tk.Text(log_frame, height=15, wrap=tk.WORD,  # Reduzido altura para caber melhor
                               font=('Consolas', 11), bg='#2c3e50', fg='#ecf0f1')  # Reduzido fonte
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=(0, 10))  # Reduzido espaçamento
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Clear log button
        clear_log_btn = ttk.Button(log_frame, text="Limpar Log", 
                                  command=self.clear_log, style='Info.TButton')
        clear_log_btn.grid(row=1, column=0, pady=(10, 0), sticky="w")  # Reduzido espaçamento
        

        
    def check_files(self):
        """Check if required files exist and update status"""
        # Check Plan.xlsx
        if os.path.exists('Plan.xlsx'):
            self.plan_status.config(text="✅ Encontrado", foreground='green')
        else:
            self.plan_status.config(text="❌ Não encontrado", foreground='red')
            
        # Check Yazo.xlsx
        if os.path.exists('Yazo.xlsx'):
            self.yazo_status.config(text="✅ Encontrado", foreground='green')
        else:
            self.yazo_status.config(text="❌ Não encontrado", foreground='red')
            
        # Only log if log_text exists (to avoid errors during initialization)
        if hasattr(self, 'log_text'):
            self.log_message("Verificação de arquivos concluída")
        
    def log_message(self, message):
        """Add message to log with timestamp"""
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def clear_log(self):
        """Clear the log text area"""
        self.log_text.delete(1.0, tk.END)
        

        
    def transfer_data(self, tipo_evento):
        """Transfer data in a separate thread to avoid GUI freezing"""
        if tipo_evento in self.opcoes_utilizadas:
            result = messagebox.askyesno("Confirmação", 
                                       f"Você já utilizou esta opção nesta sessão. Deseja prosseguir?")
            if not result:
                return
                
        # Disable buttons during transfer
        self.disable_buttons()
        
        # Run transfer in separate thread
        thread = threading.Thread(target=self._transfer_data_thread, args=(tipo_evento,))
        thread.daemon = True
        thread.start()
        
    def _transfer_data_thread(self, tipo_evento):
        """Thread function for data transfer"""
        try:
            # Mapeamento de tipos de evento para nomes de abas
            mapeamento_abas = {
                '1': 'PALESTRA',
                '2': 'WORKSHOP', 
                '3': 'PAINEL'
            }
            
            aba_nome = mapeamento_abas.get(tipo_evento)
            self.log_message(f"Iniciando transferência para aba: {aba_nome}")
            
            # Extract data from Plan.xlsx
            linhas_extraidas = self.extrair_dados_plan(aba_nome)
            
            if not linhas_extraidas:
                self.log_message(f"Nenhuma linha com fundo verde encontrada na aba {aba_nome}!")
                self.root.after(0, lambda: messagebox.showwarning("Aviso", 
                    f"Nenhuma linha com fundo verde encontrada na aba {aba_nome}!"))
                return
                
            self.log_message(f"Extraídas {len(linhas_extraidas)} linhas da aba {aba_nome}.")
            
            # Transfer to Yazo.xlsx
            if self.transferir_para_yazo(linhas_extraidas, aba_nome):
                self.opcoes_utilizadas.add(tipo_evento)
                self.log_message("Transferência concluída com sucesso!")
                self.root.after(0, lambda: messagebox.showinfo("Sucesso", 
                    f"Transferência concluída! {len(linhas_extraidas)} registros adicionados."))
            else:
                self.log_message("Erro na transferência!")
                self.root.after(0, lambda: messagebox.showerror("Erro", "Erro na transferência!"))
                
        except Exception as e:
            self.log_message(f"Erro: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro: {str(e)}"))
        finally:
            self.root.after(0, self.enable_buttons)
            
    def clear_data(self):
        """Clear data with confirmation"""
        if '5' in self.opcoes_utilizadas:
            result = messagebox.askyesno("Confirmação", 
                                       "Você já utilizou esta opção nesta sessão. Deseja prosseguir?")
            if not result:
                return
                
        result = messagebox.askyesno("Confirmação", 
                                   "ATENÇÃO: Esta ação irá APAGAR TODOS os registros da aba Palestras2k25. Deseja prosseguir?")
        if result:
            self.disable_buttons()
            
            thread = threading.Thread(target=self._clear_data_thread)
            thread.daemon = True
            thread.start()
            
    def _clear_data_thread(self):
        """Thread function for clearing data"""
        try:
            if self.apagar_dados_yazo():
                self.opcoes_utilizadas.add('5')
                self.log_message("Dados apagados com sucesso!")
                self.root.after(0, lambda: messagebox.showinfo("Sucesso", "Dados apagados com sucesso!"))
            else:
                self.log_message("Erro ao apagar dados!")
                self.root.after(0, lambda: messagebox.showerror("Erro", "Erro ao apagar dados!"))
        except Exception as e:
            self.log_message(f"Erro: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro: {str(e)}"))
        finally:
            self.root.after(0, self.enable_buttons)
            
    def disable_buttons(self):
        """Disable all buttons during operations"""
        self.palestra_btn.config(state='disabled')
        self.workshop_btn.config(state='disabled')
        self.painel_btn.config(state='disabled')
        self.clear_btn.config(state='disabled')
        
    def enable_buttons(self):
        """Enable all buttons after operations"""
        self.palestra_btn.config(state='normal')
        self.workshop_btn.config(state='normal')
        self.painel_btn.config(state='normal')
        self.clear_btn.config(state='normal')

    # ============================================================================
    # FUNÇÕES DE PROCESSAMENTO DE DADOS DO EXCEL
    # ============================================================================
    
    def verificar_celula_verde(self, cell):
        # Check if cell has the specific green background (RGB 0,255,0 / #00FF00)
        if cell.fill.start_color.rgb:
            rgb = cell.fill.start_color.rgb
            # Remove 'FF' prefix if present (Excel sometimes adds alpha channel)
            if rgb.startswith('FF'):
                rgb = rgb[2:]
            
            # Check for the specific green color used in the Excel filter
            if rgb.upper() == '00FF00':  # RGB(0,255,0) - Pure green
                return True
        
        return False

    def verificar_linha_verde(self, worksheet, row_num):
        # Check if column A has green background
        cell = worksheet.cell(row=row_num, column=1)
        return self.verificar_celula_verde(cell)

    def verificar_linha_verde_flexivel(self, worksheet, row_num):
        """Check if column A has green background"""
        # Check only column A (first column)
        cell = worksheet.cell(row=row_num, column=1)
        return self.verificar_celula_verde(cell)

    def extrair_dados_plan(self, aba_nome):
        """Extract data from Plan.xlsx based on green background cells"""
        self.log_message(f'Lendo Plan.xlsx - aba {aba_nome}...')
        
        # Load workbook and worksheet
        plan_wb = load_workbook('Plan.xlsx')
        if aba_nome not in plan_wb.sheetnames:
            self.log_message(f'Aba {aba_nome} não encontrada em Plan.xlsx!')
            return []
        
        worksheet = plan_wb[aba_nome]
        linhas_extraidas = []
        linhas_verdes_encontradas = 0
        linhas_processadas = 0
        linhas_ignoradas = []
        
        # Process each row
        for row_num in range(2, worksheet.max_row + 1):  # Skip header row
            linhas_processadas += 1
            
            # Check if column A has green background
            cell_a = worksheet.cell(row=row_num, column=1)
            is_green = self.verificar_celula_verde(cell_a)
            
            if is_green:
                linhas_verdes_encontradas += 1
                
                # Extract data from specific columns based on mapping
                nome = worksheet.cell(row=row_num, column=1).value  # A
                local = worksheet.cell(row=row_num, column=14).value  # N
                titulo = worksheet.cell(row=row_num, column=6).value  # F
                descritivo = worksheet.cell(row=row_num, column=7).value  # G
                data = worksheet.cell(row=row_num, column=11).value  # K
                hora_inicio = worksheet.cell(row=row_num, column=12).value  # L
                hora_fim = worksheet.cell(row=row_num, column=13).value  # M
                
                # Check if we have at least a name or title
                if not nome and not titulo:
                    linhas_ignoradas.append(f"Linha {row_num}: Sem nome e título")
                    continue
                
                # Format data
                if data:
                    try:
                        data_formatada = pd.to_datetime(str(data)).strftime('%Y/%m/%d')
                    except:
                        data_formatada = str(data) if data else ''
                else:
                    data_formatada = ''
                
                # Format time
                hora_inicio_str = str(hora_inicio) if hora_inicio else ''
                hora_fim_str = str(hora_fim) if hora_fim else ''
                
                # Determine tag based on aba
                tag = aba_nome.upper()
                
                linhas_extraidas.append({
                    'ID Externo': '',
                    'Local': str(local) if local else 'A ser anunciado',
                    'Local (Inglês)': '',
                    'Local (Espanhol)': '',
                    'Título*': str(titulo) if titulo else '',
                    'Título (Inglês)': '',
                    'Título (Espanhol)': '',
                    'Descrição': str(descritivo) if descritivo else '',
                    'Descrição (Inglês)': '',
                    'Descrição (Espanhol)': '',
                    'Data (YYYY/MM/DD)*': data_formatada,
                    'Horário de início (hh:mm:ss)*': hora_inicio_str,
                    'Horário de término (hh:mm:ss)*': hora_fim_str,
                    'Nome dos Palestrantes (separados por vírgula)': str(nome) if nome else '',
                    'Tags (separadas por vírgulas)': tag,
                    'Palavras-chave (separadas por vírgulas)': ''
                })
            else:
                # Debug: Check what color this cell actually has
                if cell_a.fill.start_color.rgb:
                    rgb = cell_a.fill.start_color.rgb
                    if rgb.startswith('FF'):
                        rgb = rgb[2:]
                    linhas_ignoradas.append(f"Linha {row_num}: RGB={rgb}, Theme={cell_a.fill.start_color.theme}")
                elif cell_a.fill.fill_type and cell_a.fill.fill_type != 'none':
                    linhas_ignoradas.append(f"Linha {row_num}: Fill type={cell_a.fill.fill_type}")
                else:
                    linhas_ignoradas.append(f"Linha {row_num}: Sem preenchimento")
        
        self.log_message(f'Total de linhas processadas: {linhas_processadas}')
        self.log_message(f'Linhas com coluna A verde encontradas: {linhas_verdes_encontradas}')
        self.log_message(f'Linhas extraídas: {len(linhas_extraidas)}')
        
        # Show first few ignored lines for debugging
        if linhas_ignoradas:
            self.log_message(f'Primeiras 5 linhas ignoradas:')
            for i, linha in enumerate(linhas_ignoradas[:5]):
                self.log_message(f'  {linha}')
            if len(linhas_ignoradas) > 5:
                self.log_message(f'  ... e mais {len(linhas_ignoradas) - 5} linhas')
        
        return linhas_extraidas

    def transferir_para_yazo(self, linhas_extraidas, aba_nome):
        """Transfer data to Yazo.xlsx based on event type"""
        # Load Yazo.xlsx
        try:
            yazo_wb = load_workbook('Yazo.xlsx')
        except FileNotFoundError:
            self.log_message('Arquivo Yazo.xlsx não encontrado!')
            return False
        except PermissionError:
            self.log_message('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
            return False
        
        # Define columns for Yazo
        colunas_yazo = [
            'ID Externo', 'Local', 'Local (Inglês)', 'Local (Espanhol)', 'Título*', 'Título (Inglês)', 'Título (Espanhol)',
            'Descrição', 'Descrição (Inglês)', 'Descrição (Espanhol)', 'Data (YYYY/MM/DD)*', 'Horário de início (hh:mm:ss)*',
            'Horário de término (hh:mm:ss)*', 'Nome dos Palestrantes (separados por vírgula)', 'Tags (separadas por vírgulas)',
            'Palavras-chave (separadas por vírgulas)'
        ]
        
        # Try to find the correct sheet name - use Palestras2k25 as default
        target_sheet = None
        for sheet_name in yazo_wb.sheetnames:
            if 'Palestras2k25' in sheet_name or '2025' in sheet_name or sheet_name.strip() == '2025':
                target_sheet = sheet_name
                break
        
        # Check if target sheet exists, create if not
        if not target_sheet:
            target_sheet = 'Palestras2k25'
            yazo_wb.create_sheet(target_sheet)
            # Add headers
            ws = yazo_wb[target_sheet]
            for col, header in enumerate(colunas_yazo, 1):
                ws.cell(row=1, column=col, value=header)
        
        self.log_message(f'Usando aba: "{target_sheet}"')
        
        # Read existing data
        try:
            yazo_df = pd.read_excel('Yazo.xlsx', sheet_name=target_sheet)
        except:
            # Create empty DataFrame with correct columns
            yazo_df = pd.DataFrame({col: pd.Series(dtype='str') for col in colunas_yazo})
        
        # Add new rows
        novo_df = pd.DataFrame(linhas_extraidas)
        resultado_df = pd.concat([yazo_df, novo_df], ignore_index=True)
        
        # Save back to Yazo.xlsx
        try:
            with pd.ExcelWriter('Yazo.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                resultado_df.to_excel(writer, sheet_name=target_sheet, index=False)
        except PermissionError:
            self.log_message('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
            return False
        
        self.log_message(f'Transferência concluída! {len(linhas_extraidas)} registros adicionados à aba "{target_sheet}".')
        return True

    def listar_abas_yazo(self):
        """List all available sheets in Yazo.xlsx"""
        try:
            yazo_wb = load_workbook('Yazo.xlsx')
            self.log_message("Abas disponíveis em Yazo.xlsx:")
            for i, sheet_name in enumerate(yazo_wb.sheetnames, 1):
                self.log_message(f"  {i}. '{sheet_name}'")
            return yazo_wb.sheetnames
        except FileNotFoundError:
            self.log_message('Arquivo Yazo.xlsx não encontrado!')
            return []
        except PermissionError:
            self.log_message('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
            return []

    def apagar_dados_yazo(self):
        """Clear all data below header in Yazo.xlsx sheet"""
        try:
            yazo_wb = load_workbook('Yazo.xlsx')
        except FileNotFoundError:
            self.log_message('Arquivo Yazo.xlsx não encontrado!')
            return False
        except PermissionError:
            self.log_message('ERRO: O arquivo Yazo.xlsx está aberto no Excel. Feche o arquivo e tente novamente.')
            return False
        
        # List available sheets first
        self.log_message("Abas disponíveis em Yazo.xlsx:")
        for sheet_name in yazo_wb.sheetnames:
            self.log_message(f"  - '{sheet_name}'")
        
        # Try to find the correct sheet name - use Palestras2k25 as default
        target_sheet = None
        for sheet_name in yazo_wb.sheetnames:
            if 'Palestras2k25' in sheet_name or '2025' in sheet_name or sheet_name.strip() == '2025':
                target_sheet = sheet_name
                break
        
        if not target_sheet:
            self.log_message('Aba Palestras2k25 ou 2025 não encontrada em Yazo.xlsx!')
            self.log_message('Abas disponíveis: ' + str(yazo_wb.sheetnames))
            return False
        
        self.log_message(f'Usando aba: "{target_sheet}"')
        
        # Get the worksheet
        ws = yazo_wb[target_sheet]
        
        # Clear all rows except header (row 1)
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)
        
        # Save the workbook
        yazo_wb.save('Yazo.xlsx')
        self.log_message(f'Todos os dados da aba "{target_sheet}" foram apagados (mantendo apenas o cabeçalho).')
        return True

# ============================================================================
# FUNÇÃO PRINCIPAL E INICIALIZAÇÃO
# ============================================================================

def main():
    root = tk.Tk()
    app = DataTransferApp(root)
    root.mainloop()

if __name__ == "__main__":
    main() 